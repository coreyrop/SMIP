from pyspim.pyspim import Spim
from gui.Utilities import get_path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from multiprocessing.pool import ThreadPool
from multiprocessing import cpu_count


def run_MIPS(filename):
    sp = Spim(debug=False)
    sp.load(filename)
    sp.run()
    return {i: sp.reg(i) for i in range(32)}


def grade_submissions(directory_path, lesson):
    from os import listdir
    global_path = get_path(directory_path)
    suffix = '_'+lesson.lesson_title+'(Submission).s'
    submissions = [(submission[:submission.find(suffix)], global_path+'/'+submission) for submission in listdir(global_path) if submission.endswith(suffix)]
    results = check_all_submissions(submissions, lesson)
    create_grade_workbook(lesson, results)


def create_grade_workbook(lesson, results):
    filename = get_path('/grading/'+lesson.lesson_title+'(Grades).xlsx')
    book = Workbook()
    sheet = book.active

    greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    cyanFill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sheet.column_dimensions['A'].width = 50

    def get_sheet_column(reg_num, recursed=False):
        from string import ascii_uppercase

        if reg_num < 24:
            if recursed:
                return ascii_uppercase[reg_num]
            else:
                return ascii_uppercase[reg_num+2]
        else:
            return 'A' + get_sheet_column(reg_num-24, True)

    sheet['A1'] = 'Name'
    sheet['B1'] = 'Pass/Fail'
    for i in range(32):
        index = get_sheet_column(i)+'1'
        sheet[index] = '$r'+str(i)

    row = 2
    for sub in results:
        sheet['A'+str(row)] = sub['name']

        sheet['B'+str(row)] = 'Passed' if sub['pass'] else 'Failed'
        sheet['B' + str(row)].fill = greenFill if sub['pass'] else redFill

        for i in range(32):
            index = get_sheet_column(i)+str(row)
            sheet[index] = sub[i][1]
            check = sub[i][0]
            if check is None:
                sheet[index].fill = cyanFill
            elif check:
                sheet[index].fill = greenFill
            else:
                sheet[index].fill = redFill
        row += 1
    book.save(filename)
    pass


def check_all_submissions(submissions, lesson):
    def get_results(submission):
        name, file = submission
        result = {'name': name}
        solution = run_MIPS(file)

        passed = True
        for answer_register, answer_value in lesson.lesson_answer.items():
            solution_value = solution.get(answer_register, None)
            if answer_value != solution_value:
                passed = False
                result[answer_register] = (False, solution_value)
            else:
                result[answer_register] = (True, solution_value)

        result['pass'] = passed

        for reg, val in solution.items():
            if reg not in lesson.lesson_answer.keys():
                result[reg] = (None, val)

        return result
    pool = ThreadPool(cpu_count())
    results = pool.map(get_results, submissions)
    pool.close()
    return results
