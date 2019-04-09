from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from .Lesson_Transition import dict
from .Lesson import Lesson
from gui.Utilities import get_path
import re
import os
from pathlib import Path


def get_register_index(register_num, get_value_index=True):
    if get_value_index:
        if register_num < 16:
            return 'D' + str(register_num+1)
        else:
            return 'F' + str(register_num+1 - 16)
    else:
        if register_num < 16:
            return 'C' + str(register_num+1)
        else:
            return 'E' + str(register_num+1 - 16)


def initialize_workbook(filename='temp', **kwargs):
    book = Workbook()
    sheet = book.active

    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 50
    sheet.column_dimensions['I'].width = 50

    greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    cyanFill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    labelFont = Font(name='Calibri', size=14, bold=True)
    valueFont = Font(name='Calibri', size=14)

    sheet['A1'] = 'Lesson Title'
    sheet['A2'] = 'Lesson Prompt'
    sheet['A3'] = 'Lesson Hint'
    sheet['A4'] = 'Base Code Relative File Path'
    sheet['B1'] = kwargs['lesson_title']
    sheet['B2'] = kwargs['lesson_prompt']
    sheet['B3'] = kwargs['lesson_hint']
    sheet['B4'] = kwargs['lesson_filepath']

    sheet['A1'].fill = greenFill
    sheet['A2'].fill = greenFill
    sheet['A3'].fill = greenFill
    sheet['A4'].fill = greenFill
    sheet['A1'].font = labelFont
    sheet['A2'].font = labelFont
    sheet['A3'].font = labelFont
    sheet['A4'].font = labelFont

    sheet['B1'].fill = cyanFill
    sheet['B2'].fill = cyanFill
    sheet['B3'].fill = cyanFill
    sheet['B4'].fill = cyanFill
    sheet['B1'].font = valueFont
    sheet['B2'].font = valueFont
    sheet['B3'].font = valueFont
    sheet['B4'].font = valueFont

    for i in range(32):
        sheet[get_register_index(i, False)] = '$r'+str(i)
        sheet[get_register_index(i, False)].fill = greenFill
        sheet[get_register_index(i, False)].font = labelFont

        sheet[get_register_index(i)] = kwargs['registers'][i]
        sheet[get_register_index(i)].fill = cyanFill
        sheet[get_register_index(i)].font = valueFont

    sheet['G1'] = 'Reference Name'
    sheet['H1'] = 'Reference Type'
    sheet['I1'] = 'Reference Path'

    sheet['G1'].fill = greenFill
    sheet['H1'].fill = greenFill
    sheet['I1'].fill = greenFill
    sheet['G1'].font = labelFont
    sheet['H1'].font = labelFont
    sheet['I1'].font = labelFont

    index = 2
    for reference in kwargs.get('references', []):
        sheet['G'+str(index)] = reference['Name']
        sheet['H'+str(index)] = reference['Type']
        sheet['I'+str(index)] = reference['Path']
        if reference['Type'] == 'web_link':
            sheet['I'+str(index)].style = 'Hyperlink'

        sheet['G' + str(index)].fill = cyanFill
        sheet['H' + str(index)].fill = cyanFill
        sheet['I' + str(index)].fill = cyanFill
        sheet['G' + str(index)].font = valueFont
        sheet['H' + str(index)].font = valueFont
        sheet['I' + str(index)].font = valueFont
        index += 1

    book.save(filename+'.xlsx')
    return load_lesson_from_workbook(filename+'.xlsx')


def load_lesson_from_workbook(filename):
    book = load_workbook(filename)
    sheet = book.active
    answer = {i: int(sheet[get_register_index(i)].value) for i in range(32) if
                sheet[get_register_index(i)].value is not None and re.match('[+-]?\d', sheet[
                    get_register_index(i)].value) is not None}

    references = []
    index = 2
    while sheet['G'+str(index)].value is not None:
        references.append({'Name': sheet['G'+str(index)].value, 'Type': sheet['H'+str(index)].value, 'Path': sheet['I'+str(index)].value})
        index += 1
    return Lesson(sheet['B1'].value, sheet['B2'].value, answer, [sheet['B3'].value], references, sheet['B4'].value, dict.get(sheet['B1'].value, False))


def load_lessons():
    lesson_path = get_path('/lesson_files/')
    return [load_lesson_from_workbook(lesson_path + file) for file in os.listdir(lesson_path) if file.endswith('.xlsx')]

